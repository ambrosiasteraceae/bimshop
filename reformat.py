import openpyxl
import pandas as pd
import os
import numpy as np
from openpyxl.styles import PatternFill

def parsedf():


    cd = os.getcwd()
    os.path.join(cd, '')

    ffp = 'IntermediateSchedule.csv'
    df = pd.read_csv(ffp)
    df.columns

    col_mask = [
        'Task Nesting',
        'Name',
        'Display ID',
        'Planned Start',
        'Planned End',
        'Actual Start',
        'Actual End',
        'Task Type',
        'User 1',
        'User 2'
    ]
    # df.drop_duplicates(inplace = True)

    df = df[col_mask]

    date_format = "%d/%m/%Y"

    mydates = ['Planned Start', 'Planned End', 'Actual Start', 'Actual End']
    for date in mydates:
        df[date] = pd.to_datetime(df[date], format=date_format, errors='ignore')

    def reorder_df():
        # Orders the df based on the slab order
        ordered_slabs = ['SL2-1', 'SL1-1', 'SL1-2', 'SL1-3', 'SL3-1', 'SL2-2', 'SL1-4', 'SL1-5', 'SL1-6', 'SL3-2',
                         'SL2-3', 'SL1-7', 'SL1-8', 'SL1-9', 'SL3-3', 'SL2-4', 'SL1-10', 'SL1-11', 'SL1-12', 'SL3-4',
                         'SL2-5', 'SL1-13', 'SL1-14', 'SL1-15', 'SL3-5', 'SL2-6', 'SL1-16', 'SL1-17', 'SL1-18', 'SL3-6',
                         'SL4A-1', 'SL5A-1', 'SL6A-1', 'SL7A-1', 'SL8A-1', 'SL4B-1', 'SL5B-1', 'SL6B-1', 'SL7B-1',
                         'SL8B-1', 'SL2-7', 'SL1-19', 'SL1-20', 'SL1-21', 'SL3-7', 'SL2-8', 'SL1-22', 'SL1-23',
                         'SL1-24', 'SL3-8', 'SL2-9', 'SL1-25', 'SL1-26', 'SL1-27', 'SL3-9', 'SL2-10', 'SL1-28',
                         'SL1-29', 'SL1-30', 'SL3-10', 'SL2-11', 'SL1-31', 'SL1-32', 'SL1-33', 'SL3-11', 'SL2-12',
                         'SL1-34', 'SL1-35', 'SL1-36', 'SL3-12', 'SL2-13', 'SL1-37', 'SL1-38', 'SL1-39', 'SL3-13',
                         'SL2-14', 'SL1-40', 'SL1-41', 'SL1-42', 'SL3-14', 'SL2-15', 'SL1-43', 'SL1-44', 'SL1-45',
                         'SL3-15', 'SL2-16', 'SL1-46', 'SL1-47', 'SL1-48', 'SL3-16', 'SL2-17', 'SL1-49', 'SL1-50',
                         'SL1-51', 'SL3-17', 'SL2-18', 'SL1-52', 'SL1-53', 'SL1-54', 'SL3-18', 'SL2-19', 'SL1-55',
                         'SL1-56', 'SL1-57', 'SL3-19', 'SL2-20', 'SL1-58', 'SL1-59', 'SL1-60', 'SL3-20', 'SL2-21',
                         'SL1-61', 'SL1-62', 'SL1-63', 'SL3-21', 'SL4A-2', 'SL5A-2', 'SL6A-2', 'SL7A-2', 'SL8A-2',
                         'SL8A-2', 'SL4B-2', 'SL5B-2', 'SL6B-2', 'SL7B-2', 'SL8B-2', 'SL2-22', 'SL1-64', 'SL1-65',
                         'SL1-66', 'SL3-22', 'SL2-23', 'SL1-67', 'SL1-68', 'SL1-69', 'SL3-23', 'SL2-24', 'SL1-70',
                         'SL1-71', 'SL1-72', 'SL3-24', 'SL2-25', 'SL1-73', 'SL1-74', 'SL1-75', 'SL3-25', 'SL2-26',
                         'SL1-76', 'SL1-77', 'SL1-78', 'SL3-26', 'SL2-27', 'SL1-79', 'SL1-80', 'SL1-81', 'SL3-27',
                         'SL2-28', 'SL1-82', 'SL1-83', 'SL1-84', 'SL3-28']

        # Create a mask to filter rows with matching names
        mask = df['Name'].apply(lambda x: x.strip() in ordered_slabs)
        # Create a DataFrame with matching rows and reorder them
        matching_rows = df[mask]

        initial_idx = matching_rows.index.to_list()
        myidx = initial_idx.copy()

        matching_rows['order'] = matching_rows['Name'].apply(lambda x: ordered_slabs.index(x))
        matching_rows = matching_rows.sort_values(by='order')
        matching_rows.index = myidx

        for index in matching_rows.index:
            df.loc[index] = matching_rows.loc[index]

    reorder_df()
    df.reset_index(inplace=True, drop=True)

    spacer = {
        0: "",
        1: " " * 2,
        2: " " * 6,
        3: " " * 10,
        4: " " * 14,
    }

    def add_tabs(row):
        tabs = spacer.get(row['Task Nesting'], spacer[4])
        return tabs + row['Name']

    df['Name'] = df.apply(add_tabs, axis=1)

    df['Duration'] = df['Planned End'] - df['Planned Start']

    idx = df['Duration'][df['Duration'].isna()].index.to_numpy()

    diff = idx[:-1] - idx[1:]
    diff = np.insert(diff, 0, 0)
    mask = idx[(diff < -1) | (diff == 0)]

    mask_above = mask - 1
    # mask_below = mask + 1
    df['Name'][mask_above]
    lists = np.split(idx, np.where(np.diff(idx) != 1)[0] + 1)
    from collections import namedtuple

    Task = namedtuple('Task', ['name', 'start', 'end', 'index', 'duration', 'start_days', 'end_days'])
    tasks = []
    for i, arr in zip(mask_above, lists):
        tasks.append(Task(
            df['Name'][i],
            df['Planned Start'][i],
            df['Planned End'][i],
            arr,
            df['Duration'][i],
            [],
            []
        ))

    for t in tasks:
        delta = t.duration / t.index.size
        enddate = t.start + delta
        startdate = t.start
        for i in range(t.index.size):
            t.end_days.append(enddate)
            t.start_days.append(startdate)
            startdate += delta
            enddate += delta

    def fetch_days(tasks: list) -> (list, list):
        """
        """
        a = []
        b = []
        c = []
        for task in tasks:
            a += task.index.tolist()
            b += task.start_days
            c += task.end_days
        return a, b, c

    indexes, sdays, edays = fetch_days(tasks)
    df['Planned Start'][indexes] = sdays
    df['Planned End'][indexes] = edays

    date_format = "%d/%m/%Y"
    date_format_user = "%d/%m/%Y/%H"
    for col in mydates:
        if col in ['Planned Start', 'Planned End']:
            df[col] = df[col].dt.strftime(date_format_user)
        else:
            df[col] = df[col].dt.strftime(date_format)

    df.drop(['Duration'], axis=1, inplace=True)
    df.to_excel('FinalSchedule.xlsx', index=False, )

def formatdf():
    wk = openpyxl.load_workbook('FinalSchedule.xlsx')
    print(type(wk))
    ws = wk['Sheet1']

    erange = ws.dimensions
    print(erange)

    def get_patterns():

        from openpyxl.styles.colors import Color
        # colors = (Color(indexed=32), Color(indexed=33), Color(indexed=34), Color(indexed=35))
        patterns = []
        colors = ('80765d69', '80f1828d', '80fcd0ba', '80fefad4')

        for color in colors:
            pattern = PatternFill(start_color=color, end_color=color, fill_type='darkGrid')
            patterns.append(pattern)

        return patterns

    patterns = get_patterns()

    def apply_patterns_to_empty_cells(ws, patterns):
        headercolor = '808FB9A8'
        headfill = PatternFill(start_color=headercolor, end_color=headercolor, fill_type='darkGrid')
        for i, row in enumerate(ws[erange]):
            if i == 0:
                for cell in row:
                    cell.fill = headfill
            if not row[-1].value:
                for cell in row:
                    cell.fill = patterns[row[0].value]

    # Assuming you have already defined the 'erange' and 'patterns' variables.
    apply_patterns_to_empty_cells(ws, patterns)
    desired_row_height = 20
    desired_column_width = [9, 43, 53, 12, 12, 12, 12, 9, 9, 9]
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = desired_column_width[col[0].column - 1]
    # Loop through all rows and set the row height
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = desired_row_height
    wk.save('FinalSchedule.xlsx')

if __name__ == '__main__':
    parsedf()
    formatdf()